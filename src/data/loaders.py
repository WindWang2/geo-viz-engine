import json
import shutil
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path

from src.data.models import (
    CurveData,
    FaciesData,
    IntervalItem,
    LineStyle,
    WellCoordinates,
    WellIntervals,
    WellLogData,
)

_SS_NS = "{urn:schemas-microsoft-com:office:spreadsheet}"

_SENTINEL_VALUES = {-9999, -999.25, -9999.0, -999.25}


def load_well_coordinates(path: Path) -> list[WellCoordinates]:
    if not path.exists():
        return []
    with open(path) as f:
        raw = json.load(f)
    # Support both flat list format and {"wells": [...]} dict format
    if isinstance(raw, dict):
        items = raw.get("wells", [])
    else:
        items = raw
    result = []
    for w in items:
        # Map well_name -> name if needed
        if "well_name" in w and "name" not in w:
            w = {**w, "name": w["well_name"]}
        result.append(WellCoordinates(**w))
    return result


def _open_xls_as_xlsx(path: Path):
    """Open a .xls file that is actually xlsx format (common mislabeling)."""
    import openpyxl

    tmp = None
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True)
        return wb
    except Exception:
        pass

    # Try with .xlsx extension
    tmp = str(path) + ".xlsx"
    shutil.copy2(str(path), tmp)
    try:
        wb = openpyxl.load_workbook(tmp, read_only=True)
        return wb
    finally:
        import os
        if tmp and os.path.exists(tmp):
            os.remove(tmp)


def _read_interval_sheet(ws, name_col: int, top_col: int, bottom_col: int,
                         text_col: int | None = None) -> list[IntervalItem]:
    """Read an interval sheet with columns: name, top, bottom [, text]."""
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[top_col] is None or row[bottom_col] is None:
            continue
        name = row[text_col] if text_col is not None and row[text_col] else row[name_col]
        if not name:
            continue
        items.append(IntervalItem(top=float(row[top_col]), bottom=float(row[bottom_col]), name=str(name)))
    return items


def _read_curve_sheet(ws, col_indices: list[int], names: list[str],
                      units: list[str], colors: list[str],
                      display_ranges: list[tuple[float, float]]) -> list[CurveData]:
    """Read curve data from a sheet. Depth is always column 0."""
    depths = []
    curve_values: dict[str, list[float]] = {n: [] for n in names}

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if d is None:
            continue
        depth_val = float(d)

        # Check if all curve values are sentinel
        vals = []
        all_sentinel = True
        for idx in col_indices:
            v = row[idx] if idx < len(row) else None
            if v is not None and v not in _SENTINEL_VALUES:
                all_sentinel = False
                vals.append(float(v))
            else:
                vals.append(None)

        if all_sentinel:
            continue

        depths.append(depth_val)
        for i, name in enumerate(names):
            curve_values[name].append(vals[i] if vals[i] is not None else float("nan"))

    curves = []
    for i, name in enumerate(names):
        if not curve_values[name]:
            continue
        curves.append(CurveData(
            name=name,
            unit=units[i],
            depth=depths[:],
            values=curve_values[name],
            display_range=display_ranges[i],
            color=colors[i],
        ))
    return curves


def _read_facies_from_xml(path: Path) -> FaciesData:
    """Read facies data from Excel XML Spreadsheet 2003 format."""
    facies = FaciesData()
    if not path.exists():
        return facies

    tree = ET.parse(str(path))
    root = tree.getroot()

    # Find the 文本道 worksheet
    for ws in root.iter(f"{_SS_NS}Worksheet"):
        name = ws.get(f"{_SS_NS}Name", "")
        if name != "文本道":
            continue

        table = ws.find(f"{_SS_NS}Table")
        if table is None:
            break

        rows = table.findall(f"{_SS_NS}Row")
        for row in rows[1:]:  # skip header
            cells = row.findall(f"{_SS_NS}Cell")
            # Extract cell values, handling merged/missing cells via index attribute
            values: list[str | None] = []
            next_idx = 0
            for cell in cells:
                idx_attr = cell.get(f"{_SS_NS}Index")
                if idx_attr:
                    cur_idx = int(idx_attr) - 1  # 1-based to 0-based
                    while len(values) < cur_idx:
                        values.append(None)
                data_el = cell.find(f"{_SS_NS}Data")
                values.append(data_el.text if data_el is not None and data_el.text else None)

            if len(values) < 10:
                continue
            track_name = values[1] or ""
            top_str = values[3]
            bottom_str = values[6]
            name_str = values[9]
            if not top_str or not bottom_str or not name_str:
                continue

            item = IntervalItem(top=float(top_str), bottom=float(bottom_str), name=str(name_str))
            if "微相" in track_name:
                facies.micro_phase.append(item)
            elif "亚相" in track_name:
                facies.sub_phase.append(item)
            elif "相" in track_name:
                facies.phase.append(item)
        break

    return facies


def load_well_log_from_excel(path: Path) -> WellLogData:
    wb = _open_xls_as_xlsx(path)

    # --- Curves ---
    curves: list[CurveData] = []

    # GR sheet: depth, GR
    if "GR" in wb.sheetnames:
        ws = wb["GR"]
        curves.extend(_read_curve_sheet(
            ws, col_indices=[1], names=["GR"], units=["gAPI"],
            colors=["#63b3ed"], display_ranges=[(0, 150)],
        ))

    # 电阻率 sheet: depth, R39AC, R15PC, R27PC, R39PC
    if "电阻率" in wb.sheetnames:
        ws = wb["电阻率"]
        curves.extend(_read_curve_sheet(
            ws, col_indices=[1, 2, 3, 4],
            names=["R39AC", "R15PC", "R27PC", "R39PC"],
            units=["Ω·m", "Ω·m", "Ω·m", "Ω·m"],
            colors=["#f6ad55", "#ed8936", "#dd6b20", "#c05621"],
            display_ranges=[(0.1, 50), (0.01, 10), (0.1, 50), (0.1, 100)],
        ))

    # 孔隙度 sheet: depth, 孔隙度, 渗透率, 含水饱和度, RHOB, TNPH, PE
    if "孔隙度" in wb.sheetnames:
        ws = wb["孔隙度"]
        curves.extend(_read_curve_sheet(
            ws, col_indices=[1, 2, 3, 4, 5, 6],
            names=["PHIE", "PERM", "SW", "RHOB", "TNPH", "PE"],
            units=["%", "mD", "%", "g/cm³", "dec", "b/e"],
            colors=["#68d391", "#48bb78", "#38a169", "#9f7aea", "#805ad5", "#6b46c1"],
            display_ranges=[(0, 40), (0.01, 10000), (0, 100), (1.9, 3.0), (0, 0.6), (0, 8)],
        ))

    # --- Lithology ---
    lithology: list[IntervalItem] = []
    if "岩性剖面" in wb.sheetnames:
        lithology = _read_interval_sheet(wb["岩性剖面"], name_col=0, top_col=1, bottom_col=2, text_col=3)

    # --- Stratigraphy ---
    series: list[IntervalItem] = []
    formation: list[IntervalItem] = []
    if "地层系统" in wb.sheetnames:
        ws = wb["地层系统"]
        group = "series"
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] is None:
                group = "formation"
                continue
            item = IntervalItem(top=float(row[1]), bottom=float(row[2]), name=str(row[3]))
            if group == "series":
                series.append(item)
            else:
                formation.append(item)

    # --- Facies (from XML file, not in xlsx) ---
    xml_path = path.parent / "25-HZ25-10-1井单井综合柱状图-2019-沉积室-勘探室-无地化录井-测井.xml"
    facies = _read_facies_from_xml(xml_path)

    # --- Determine depth range ---
    all_depths = []
    for c in curves:
        all_depths.extend(d for d in c.depth if d == d)  # filter NaN
    top_depth = min(all_depths) if all_depths else 0
    bottom_depth = max(all_depths) if all_depths else 100

    wb.close()

    return WellLogData(
        well_name="HZ25-10-1",
        top_depth=top_depth,
        bottom_depth=bottom_depth,
        curves=curves,
        intervals=WellIntervals(
            series=series,
            formation=formation,
            lithology=lithology,
            facies=facies,
        ),
    )


# ---------------------------------------------------------------------------
# 老龙1 well loader (true .xls format via xlrd)
# ---------------------------------------------------------------------------

def _read_xlrd_intervals(ws, top_col: int, bottom_col: int, name_col: int,
                          start_row: int = 1) -> list[IntervalItem]:
    items = []
    for i in range(start_row, ws.nrows):
        row = ws.row_values(i)
        try:
            top = float(row[top_col]) if row[top_col] != "" else None
            bottom = float(row[bottom_col]) if row[bottom_col] != "" else None
        except (ValueError, TypeError):
            continue
        if top is None or bottom is None:
            continue
        name = str(row[name_col]).strip() if name_col < len(row) and row[name_col] else None
        if name:
            items.append(IntervalItem(top=top, bottom=bottom, name=name))
    return items


def _read_xlrd_curves(ws, col_indices: list[int], names: list[str],
                      units: list[str], colors: list[str],
                      line_styles: list[LineStyle],
                      display_ranges: list[tuple[float, float]]) -> list[CurveData]:
    depths: list[float] = []
    curve_values: dict[str, list[float]] = {n: [] for n in names}

    for i in range(1, ws.nrows):
        row = ws.row_values(i)
        if not row[0]:
            continue
        try:
            d = float(row[0])
        except (ValueError, TypeError):
            continue

        depths.append(d)
        for j, name in enumerate(names):
            idx = col_indices[j]
            v = row[idx] if idx < len(row) else None
            try:
                val = float(v) if v is not None and v != "" else float("nan")
                if val in _SENTINEL_VALUES:
                    val = float("nan")
            except (ValueError, TypeError):
                val = float("nan")
            curve_values[name].append(val)

    curves = []
    for i, name in enumerate(names):
        if not depths:
            continue
        curves.append(CurveData(
            name=name,
            unit=units[i],
            depth=depths[:],
            values=curve_values[name],
            display_range=display_ranges[i],
            color=colors[i],
            line_style=line_styles[i],
        ))
    return curves


def load_well_log_laolong1(path: Path) -> WellLogData:
    import xlrd as _xlrd

    wb = _xlrd.open_workbook(str(path))
    sheet_names = wb.sheet_names()

    curves: list[CurveData] = []

    # AC/GR: depth(0), AC(1), GR(2)
    if "AC、GR" in sheet_names:
        ws = wb.sheet_by_name("AC、GR")
        curves.extend(_read_xlrd_curves(
            ws, col_indices=[1, 2],
            names=["AC", "GR"], units=["us/ft", "gAPI"],
            colors=["#60a5fa", "#4ade80"],
            line_styles=[LineStyle.DASHED, LineStyle.SOLID],
            display_ranges=[(40, 80), (0, 100)],
        ))

    # RT/RXO: depth(0), RT(1), RXO(2)
    if "RT、RXO" in sheet_names:
        ws = wb.sheet_by_name("RT、RXO")
        curves.extend(_read_xlrd_curves(
            ws, col_indices=[1, 2],
            names=["RT", "RXO"], units=["Ω·m", "Ω·m"],
            colors=["#ef4444", "#f97316"],
            line_styles=[LineStyle.SOLID, LineStyle.DASHED],
            display_ranges=[(0.1, 1000), (0.1, 1000)],
        ))

    # Lithology intervals for pattern rendering
    lithology: list[IntervalItem] = []
    if "岩性剖面" in sheet_names:
        ws = wb.sheet_by_name("岩性剖面")
        lithology = _read_xlrd_intervals(ws, top_col=2, bottom_col=3, name_col=5)

    # Lithology text descriptions
    lithology_desc: list[IntervalItem] = []
    if "岩性描述" in sheet_names:
        ws = wb.sheet_by_name("岩性描述")
        lithology_desc = _read_xlrd_intervals(ws, top_col=2, bottom_col=3, name_col=5)

    # Stratigraphy: 地层系统 sheet has repeated header blocks
    # Block 1 → 系(series), Block 2 → 统(system), Block 3 → 组(formation)
    series: list[IntervalItem] = []
    system_intervals: list[IntervalItem] = []
    formation: list[IntervalItem] = []
    if "地层系统" in sheet_names:
        ws = wb.sheet_by_name("地层系统")
        block = 0
        for i in range(ws.nrows):
            row = ws.row_values(i)
            if str(row[0]) == "井号":
                block += 1
                continue
            if block > 3 or not row[0] or not row[2] or not row[3]:
                continue
            try:
                item = IntervalItem(
                    top=float(row[2]), bottom=float(row[3]),
                    name=str(row[1]).strip(),
                )
                if block == 1:
                    series.append(item)
                elif block == 2:
                    system_intervals.append(item)
                elif block == 3:
                    formation.append(item)
            except (ValueError, TypeError):
                continue

    # Facies from dedicated sheets
    micro_phase: list[IntervalItem] = []
    if "微相" in sheet_names:
        micro_phase = _read_xlrd_intervals(
            wb.sheet_by_name("微相"), top_col=2, bottom_col=3, name_col=5)

    sub_phase: list[IntervalItem] = []
    if "亚相" in sheet_names:
        sub_phase = _read_xlrd_intervals(
            wb.sheet_by_name("亚相"), top_col=2, bottom_col=3, name_col=5)

    phase: list[IntervalItem] = []
    if "相" in sheet_names:
        phase = _read_xlrd_intervals(
            wb.sheet_by_name("相"), top_col=2, bottom_col=3, name_col=5)

    # Systems tract (Block 1) and sequence (Block 2) from 层序 sheet
    systems_tract: list[IntervalItem] = []
    sequence: list[IntervalItem] = []
    if "层序" in sheet_names:
        ws = wb.sheet_by_name("层序")
        block = 0
        for i in range(ws.nrows):
            row = ws.row_values(i)
            if str(row[0]) == "井号":
                block += 1
                continue
            if not row[2] or not row[3] or not row[5]:
                continue
            try:
                item = IntervalItem(
                    top=float(row[2]), bottom=float(row[3]),
                    name=str(row[5]).strip(),
                )
                if block == 1:
                    systems_tract.append(item)
                elif block == 2:
                    sequence.append(item)
            except (ValueError, TypeError):
                continue

    all_depths = [d for c in curves for d in c.depth if d == d]
    top_depth = min(all_depths) if all_depths else 2515.0
    bottom_depth = max(all_depths) if all_depths else 2610.0

    return WellLogData(
        well_name="老龙1",
        top_depth=top_depth,
        bottom_depth=bottom_depth,
        curves=curves,
        intervals=WellIntervals(
            series=series,
            system=system_intervals,
            formation=formation,
            lithology=lithology,
            lithology_desc=lithology_desc,
            systems_tract=systems_tract,
            sequence=sequence,
            facies=FaciesData(
                phase=phase,
                sub_phase=sub_phase,
                micro_phase=micro_phase,
            ),
        ),
    )
