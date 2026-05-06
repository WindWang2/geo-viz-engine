import os
import xml.etree.ElementTree as ET
import openpyxl

def convert_xml_to_xlsx(xml_path, xlsx_path):
    print(f"Converting {os.path.basename(xml_path)} ...")
    # Register namespaces
    ns = {
        'ss': 'urn:schemas-microsoft-com:office:spreadsheet',
        'x': 'urn:schemas-microsoft-com:office:excel',
        'o': 'urn:schemas-microsoft-com:office:office',
    }
    
    try:
        # Parse XML
        tree = ET.parse(xml_path)
        root = tree.getroot()
        
        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)
        
        # Find all worksheets
        worksheets = root.findall('.//{urn:schemas-microsoft-com:office:spreadsheet}Worksheet')
        
        for ws in worksheets:
            ws_name = ws.get('{urn:schemas-microsoft-com:office:spreadsheet}Name', 'Sheet')
            ws_out = wb_out.create_sheet(title=ws_name[:31])
            
            rows = ws.findall('.//{urn:schemas-microsoft-com:office:spreadsheet}Row')
            for r_idx, r in enumerate(rows, start=1):
                c_idx = 1
                cells = r.findall('.//{urn:schemas-microsoft-com:office:spreadsheet}Cell')
                for c in cells:
                    idx_attr = c.get('{urn:schemas-microsoft-com:office:spreadsheet}Index')
                    if idx_attr:
                        c_idx = int(idx_attr)
                    
                    data_el = c.find('.//{urn:schemas-microsoft-com:office:spreadsheet}Data')
                    if data_el is not None:
                        data_val = data_el.text
                        data_type = data_el.get('{urn:schemas-microsoft-com:office:spreadsheet}Type')
                        
                        if data_val is not None:
                            if data_type in ('Number', 'Float', 'Integer'):
                                try:
                                    if '.' in data_val:
                                        data_val = float(data_val)
                                    else:
                                        data_val = int(data_val)
                                except ValueError:
                                    pass
                            ws_out.cell(row=r_idx, column=c_idx, value=data_val)
                    c_idx += 1
                    
        if not wb_out.sheetnames:
            wb_out.create_sheet(title='Sheet1')
            
        wb_out.save(xlsx_path)
        print(f"Saved: {xlsx_path}")
    except Exception as e:
        print(f"Failed to convert {xml_path}: {e}")

def main():
    src_dir = "/home/kevin/Downloads/excel"
    dst_dir = "/home/kevin/projects/geo-viz-engine/data"
    
    os.makedirs(dst_dir, exist_ok=True)
    
    files = [f for f in os.listdir(src_dir) if f.endswith(".xml")]
    print(f"Found {len(files)} xml files to convert.")
    
    for f in files:
        xml_path = os.path.join(src_dir, f)
        # Convert e.g. "abc.xml" to "abc.xlsx"
        xlsx_name = os.path.splitext(f)[0] + ".xlsx"
        xlsx_path = os.path.join(dst_dir, xlsx_name)
        convert_xml_to_xlsx(xml_path, xlsx_path)

if __name__ == '__main__':
    main()
